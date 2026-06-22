#!/usr/bin/env python3
"""
Excel conflict workbench helper for git-safe-commit.

This tool only reads exported OURS/THEIRS workbooks or writes a merge candidate.
It never stages files and never writes to the original conflicted path.
"""

import argparse
import copy
import datetime as dt
import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print(json.dumps({"ok": False, "error": "openpyxl is required"}), file=sys.stderr)
    sys.exit(1)


LUBAN_PREFIXES = ("##var", "##type", "##group", "##")
MAX_ROWS_DEFAULT = 500
MAX_COLS_DEFAULT = 80
LUBAN_HEADER_ROWS = 4
LUBAN_DATA_START_ROW = 5


def is_xlsx(path):
    return str(path).lower().endswith(".xlsx")


def fail(message):
    print(json.dumps({"ok": False, "error": message}, ensure_ascii=False), file=sys.stderr)
    sys.exit(1)


def to_json_value(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return str(value)


def display_value(value):
    if value is None:
        return ""
    return str(to_json_value(value))


def value_kind(value):
    if isinstance(value, str) and value.startswith("="):
        return "formula"
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


def is_simple_numeric(value):
    if value is None:
        return False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return False
        try:
            float(stripped)
            return True
        except ValueError:
            return False
    return False


def cell_payload(value):
    return {
        "value": to_json_value(value),
        "display": display_value(value),
        "kind": value_kind(value),
    }


def read_non_empty_cells(ws):
    cells = {}
    if ws is None:
        return cells, 0, 0
    max_row = 0
    max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cells[(cell.row, cell.column)] = cell.value
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
    return cells, max_row, max_col


def has_luban_header(cells):
    for value in cells.values():
        if isinstance(value, str) and any(value.startswith(prefix) for prefix in LUBAN_PREFIXES):
            return True
    return False


def has_formula(cells):
    for value in cells.values():
        if isinstance(value, str) and value.startswith("="):
            return True
    return False


def values_equal(left, right):
    left_cmp = str(left) if left is not None else None
    right_cmp = str(right) if right is not None else None
    return left_cmp == right_cmp


def normalized_key(value):
    if value is None:
        return ""
    return str(value).strip()


def is_blank_row(cells, row_num, max_col):
    return all(cells.get((row_num, col_num)) is None for col_num in range(1, max_col + 1))


def detect_key_column(base_cells, ours_cells, theirs_cells, max_col):
    candidates = (base_cells, ours_cells, theirs_cells)
    for col_num in range(1, max_col + 1):
        labels = [normalized_key(cells.get((row, col_num))).lower() for cells in candidates for row in (1, 2, 3, 4)]
        if "id" in labels or "主键" in labels:
            return col_num
    return None


def row_map_by_key(cells, max_row, max_col, key_col):
    rows = {}
    row_numbers = {}
    for row_num in range(LUBAN_DATA_START_ROW, max_row + 1):
        key = normalized_key(cells.get((row_num, key_col)))
        if not key:
            if is_blank_row(cells, row_num, max_col):
                continue
            key = f"__row_{row_num}"
        if key in rows:
            key = f"{key}__row_{row_num}"
        rows[key] = {col_num: cells.get((row_num, col_num)) for col_num in range(1, max_col + 1)}
        row_numbers[key] = row_num
    return rows, row_numbers


def ordered_keys(base_keys, ours_keys, theirs_keys):
    output = []
    seen = set()
    for source in (base_keys, ours_keys, theirs_keys):
        for key in source:
            if key in seen:
                continue
            seen.add(key)
            output.append(key)
    return output


def value_from_row(rows, key, col_num):
    return rows.get(key, {}).get(col_num)


def keyed_row_status(base_exists, ours_exists, theirs_exists, has_base=True):
    if not has_base:
        if ours_exists and theirs_exists:
            return "matched"
        if ours_exists:
            return "only-ours"
        return "only-theirs"
    if base_exists and ours_exists and theirs_exists:
        return "matched"
    if not base_exists:
        return "added"
    if not ours_exists:
        return "deleted-ours"
    return "deleted-theirs"


def classify_sheet(structure_mismatch, identical, luban, real_data, formula):
    if structure_mismatch:
        return "has_impact"
    if identical:
        return "identical"
    if luban or real_data or formula:
        return "has_impact"
    return "no_impact"


def build_keyed_sheet_payload(index, ws_ours, ws_theirs, max_rows, max_cols, ws_base, base_cells, ours_cells, theirs_cells, base_max_row, ours_max_row, theirs_max_row, base_max_col, ours_max_col, theirs_max_col, key_col, focus_row_keys=None):
    ours_name = ws_ours.title if ws_ours else f"sheet{index + 1}"
    theirs_name = ws_theirs.title if ws_theirs else f"sheet{index + 1}"
    base_name = ws_base.title if ws_base else ours_name
    display_name = ours_name if ours_name == theirs_name else f"{ours_name} / {theirs_name}"
    structure_mismatch = ws_ours is None or ws_theirs is None or ours_name != theirs_name or base_name != ours_name
    has_base = ws_base is not None

    used_cols = max(base_max_col, ours_max_col, theirs_max_col)
    visible_cols = min(used_cols, max_cols)
    base_rows, base_row_numbers = row_map_by_key(base_cells, base_max_row, used_cols, key_col)
    ours_rows, ours_row_numbers = row_map_by_key(ours_cells, ours_max_row, used_cols, key_col)
    theirs_rows, theirs_row_numbers = row_map_by_key(theirs_cells, theirs_max_row, used_cols, key_col)
    keys = ordered_keys(base_rows.keys(), ours_rows.keys(), theirs_rows.keys())

    diff_keys = []
    auto_merge_keys = []
    conflict_keys = []
    rows = []
    real_data = False

    header_rows = []
    for row_num in range(1, min(LUBAN_HEADER_ROWS, max_rows) + 1):
        cells = []
        row_has_diff = False
        for col_num in range(1, visible_cols + 1):
            key = (row_num, col_num)
            base_value = base_cells.get(key)
            ours_value = ours_cells.get(key)
            theirs_value = theirs_cells.get(key)
            equal = values_equal(ours_value, theirs_value)
            row_has_diff = row_has_diff or not equal
            cells.append({
                "address": f"{get_column_letter(col_num)}{row_num}",
                "column": get_column_letter(col_num),
                "row": row_num,
                "equal": equal,
                "autoMerge": False,
                "conflict": False if equal else True,
                "base": cell_payload(base_value),
                "ours": cell_payload(ours_value),
                "theirs": cell_payload(theirs_value),
            })
        header_rows.append({"row": row_num, "rowKey": "", "hasDiff": row_has_diff, "cells": cells})

    data_row_payloads = []
    visible_data_keys = []
    key_positions = {row_key: LUBAN_DATA_START_ROW + index for index, row_key in enumerate(keys)}
    for row_key in keys:
        cells = []
        row_has_diff = False
        base_exists = row_key in base_rows
        ours_exists = row_key in ours_rows
        theirs_exists = row_key in theirs_rows
        visual_index = key_positions[row_key]
        row_conflict = False
        row_auto_merge = False
        row_status = keyed_row_status(base_exists, ours_exists, theirs_exists, has_base)
        for col_num in range(1, visible_cols + 1):
            column_letter = get_column_letter(col_num)
            base_value = value_from_row(base_rows, row_key, col_num)
            ours_value = value_from_row(ours_rows, row_key, col_num)
            theirs_value = value_from_row(theirs_rows, row_key, col_num)
            base_row_num = base_row_numbers.get(row_key)
            ours_row_num = ours_row_numbers.get(row_key)
            theirs_row_num = theirs_row_numbers.get(row_key)
            if values_equal(ours_value, theirs_value):
                equal = True
                auto_merge = False
                conflict = False
            else:
                equal = False
                row_has_diff = True
                diff_keys.append((visual_index, col_num))
                if has_base:
                    ours_changed = not values_equal(base_value, ours_value)
                    theirs_changed = not values_equal(base_value, theirs_value)
                    auto_merge = (ours_changed != theirs_changed)
                    conflict = not auto_merge
                else:
                    auto_merge = False
                    conflict = True
                if auto_merge:
                    auto_merge_keys.append((visual_index, col_num))
                    row_auto_merge = True
                else:
                    conflict_keys.append((visual_index, col_num))
                    row_conflict = True
            if ours_value is not None and not is_simple_numeric(ours_value):
                real_data = True
            if theirs_value is not None and not is_simple_numeric(theirs_value):
                real_data = True
            cells.append({
                "address": f"{column_letter}{visual_index}",
                "column": column_letter,
                "row": visual_index,
                "rowKey": row_key,
                "baseRow": base_row_num,
                "oursRow": ours_row_num,
                "theirsRow": theirs_row_num,
                "baseAddress": f"{column_letter}{base_row_num}" if base_row_num else None,
                "oursAddress": f"{column_letter}{ours_row_num}" if ours_row_num else None,
                "theirsAddress": f"{column_letter}{theirs_row_num}" if theirs_row_num else None,
                "rowStatus": row_status,
                "equal": equal,
                "autoMerge": auto_merge,
                "conflict": conflict,
                "base": cell_payload(base_value),
                "ours": cell_payload(ours_value),
                "theirs": cell_payload(theirs_value),
            })
        data_row_payloads.append({
            "row": visual_index,
            "rowKey": row_key,
            "baseRow": base_row_numbers.get(row_key),
            "oursRow": ours_row_numbers.get(row_key),
            "theirsRow": theirs_row_numbers.get(row_key),
            "rowStatus": row_status,
            "hasDiff": row_has_diff,
            "hasConflict": row_conflict,
            "hasAutoMerge": row_auto_merge,
            "cells": cells,
        })

    data_row_limit = max(0, max_rows - len(header_rows))
    important_rows = [row for row in data_row_payloads if row["hasDiff"] or row["rowStatus"] != "matched"]
    context_rows = [row for row in data_row_payloads if not row["hasDiff"] and row["rowStatus"] == "matched"]
    conflict_rows = [row for row in important_rows if row["hasConflict"]]
    auto_merge_rows = [row for row in important_rows if not row["hasConflict"] and row["hasAutoMerge"]]
    other_changed_rows = [row for row in important_rows if not row["hasConflict"] and not row["hasAutoMerge"]]
    focus_row_keys = set(focus_row_keys or [])
    focus_rows = [row for row in data_row_payloads if row["rowKey"] in focus_row_keys]
    rows = []
    seen_row_keys = set()
    for row in focus_rows + conflict_rows + auto_merge_rows + other_changed_rows + context_rows:
        if row["rowKey"] in seen_row_keys:
            continue
        seen_row_keys.add(row["rowKey"])
        rows.append(row)
        if len(rows) >= data_row_limit:
            break
    visible_data_keys = [row["rowKey"] for row in rows]

    all_rows = header_rows + rows
    luban = has_luban_header(ours_cells) or has_luban_header(theirs_cells)
    formula = has_formula(ours_cells) or has_formula(theirs_cells)
    classification = classify_sheet(
        structure_mismatch,
        len(diff_keys) == 0 and not any(row["hasDiff"] for row in header_rows),
        luban,
        real_data,
        formula,
    )
    visible_rows = len(all_rows)
    used_rows = LUBAN_HEADER_ROWS + len(keys)

    return {
        "sheetIndex": index + 1,
        "name": display_name,
        "oursName": ours_name,
        "theirsName": theirs_name,
        "classification": classification,
        "structureMismatch": structure_mismatch,
        "hasLubanHeader": luban,
        "hasFormula": formula,
        "hasRealData": real_data,
        "alignment": {
            "mode": "keyed",
            "keyColumn": get_column_letter(key_col),
            "keyColumnIndex": key_col,
            "dataStartRow": LUBAN_DATA_START_ROW,
        },
        "diffCount": len(diff_keys),
        "diffCells": [f"{get_column_letter(col)}{row}" for row, col in diff_keys[:200]],
        "diffRows": sorted({row for row, _ in diff_keys}),
        "autoMergeCount": len(auto_merge_keys),
        "autoMergeCells": [f"{get_column_letter(col)}{row}" for row, col in auto_merge_keys[:200]],
        "conflictCount": len(conflict_keys),
        "conflictCells": [f"{get_column_letter(col)}{row}" for row, col in conflict_keys[:200]],
        "usedRows": used_rows,
        "usedCols": used_cols,
        "visibleRows": visible_rows,
        "visibleCols": visible_cols,
        "truncated": len(keys) > len(visible_data_keys) or used_cols > visible_cols,
        "columns": [get_column_letter(col) for col in range(1, visible_cols + 1)],
        "rows": all_rows,
    }


def build_sheet_payload(index, ws_ours, ws_theirs, max_rows, max_cols, ws_base=None, focus_row_keys=None):
    base_cells, base_max_row, base_max_col = read_non_empty_cells(ws_base)
    ours_cells, ours_max_row, ours_max_col = read_non_empty_cells(ws_ours)
    theirs_cells, theirs_max_row, theirs_max_col = read_non_empty_cells(ws_theirs)

    ours_name = ws_ours.title if ws_ours else f"sheet{index + 1}"
    theirs_name = ws_theirs.title if ws_theirs else f"sheet{index + 1}"
    base_name = ws_base.title if ws_base else ours_name
    display_name = ours_name if ours_name == theirs_name else f"{ours_name} / {theirs_name}"
    structure_mismatch = ws_ours is None or ws_theirs is None or ours_name != theirs_name or (ws_base is not None and base_name != ours_name)
    used_cols = max(base_max_col, ours_max_col, theirs_max_col)
    key_col = detect_key_column(base_cells, ours_cells, theirs_cells, used_cols)
    if key_col and (has_luban_header(base_cells) or has_luban_header(ours_cells) or has_luban_header(theirs_cells)):
        return build_keyed_sheet_payload(index, ws_ours, ws_theirs, max_rows, max_cols, ws_base, base_cells, ours_cells, theirs_cells, base_max_row, ours_max_row, theirs_max_row, base_max_col, ours_max_col, theirs_max_col, key_col, focus_row_keys)

    all_keys = set(base_cells) | set(ours_cells) | set(theirs_cells)
    diff_keys = []
    auto_merge_keys = []
    conflict_keys = []
    real_data = False
    for key in sorted(all_keys):
        base_value = base_cells.get(key)
        ours_value = ours_cells.get(key)
        theirs_value = theirs_cells.get(key)
        if values_equal(ours_value, theirs_value):
            continue
        diff_keys.append(key)
        if ws_base is not None:
            ours_changed = not values_equal(base_value, ours_value)
            theirs_changed = not values_equal(base_value, theirs_value)
            if ours_changed and theirs_changed:
                conflict_keys.append(key)
            else:
                auto_merge_keys.append(key)
        else:
            conflict_keys.append(key)
        if ours_value is not None and not is_simple_numeric(ours_value):
            real_data = True
        if theirs_value is not None and not is_simple_numeric(theirs_value):
            real_data = True

    luban = has_luban_header(ours_cells) or has_luban_header(theirs_cells)
    formula = has_formula(ours_cells) or has_formula(theirs_cells)
    classification = classify_sheet(
        structure_mismatch,
        len(diff_keys) == 0,
        luban,
        real_data,
        formula,
    )

    used_rows = max(base_max_row, ours_max_row, theirs_max_row)
    used_cols = max(base_max_col, ours_max_col, theirs_max_col)
    visible_cols = min(used_cols, max_cols)
    diff_set = set(diff_keys)
    auto_merge_set = set(auto_merge_keys)
    conflict_set = set(conflict_keys)

    def ordered_row_numbers(keys):
        return sorted({row for row, _ in keys})

    context_row_numbers = [row for row in range(1, used_rows + 1)]
    row_numbers = []
    seen_rows = set()
    for row_num in (
        ordered_row_numbers(conflict_keys)
        + ordered_row_numbers(auto_merge_keys)
        + ordered_row_numbers([key for key in diff_keys if key not in conflict_set and key not in auto_merge_set])
        + context_row_numbers
    ):
        if row_num in seen_rows:
            continue
        seen_rows.add(row_num)
        row_numbers.append(row_num)
        if len(row_numbers) >= max_rows:
            break

    rows = []
    for row_num in row_numbers:
        cells = []
        for col_num in range(1, visible_cols + 1):
            address = f"{get_column_letter(col_num)}{row_num}"
            key = (row_num, col_num)
            ours_value = ours_cells.get(key)
            theirs_value = theirs_cells.get(key)
            cells.append({
                "address": address,
                "column": get_column_letter(col_num),
                "row": row_num,
                "equal": key not in diff_set,
                "autoMerge": key in auto_merge_set,
                "conflict": key in conflict_set,
                "base": cell_payload(base_cells.get(key)),
                "ours": cell_payload(ours_value),
                "theirs": cell_payload(theirs_value),
            })
        rows.append({"row": row_num, "hasDiff": any((row_num, col) in diff_set for col in range(1, visible_cols + 1)), "cells": cells})

    return {
        "sheetIndex": index + 1,
        "name": display_name,
        "oursName": ours_name,
        "theirsName": theirs_name,
        "classification": classification,
        "structureMismatch": structure_mismatch,
        "hasLubanHeader": luban,
        "hasFormula": formula,
        "hasRealData": real_data,
        "alignment": {
            "mode": "coordinate",
        },
        "diffCount": len(diff_keys),
        "diffCells": [f"{get_column_letter(col)}{row}" for row, col in diff_keys[:200]],
        "diffRows": sorted({row for row, _ in diff_keys}),
        "autoMergeCount": len(auto_merge_keys),
        "autoMergeCells": [f"{get_column_letter(col)}{row}" for row, col in auto_merge_keys[:200]],
        "conflictCount": len(conflict_keys),
        "conflictCells": [f"{get_column_letter(col)}{row}" for row, col in conflict_keys[:200]],
        "usedRows": used_rows,
        "usedCols": used_cols,
        "visibleRows": len(rows),
        "visibleCols": visible_cols,
        "truncated": used_rows > len(rows) or used_cols > visible_cols,
        "columns": [get_column_letter(col) for col in range(1, visible_cols + 1)],
        "rows": rows,
    }


def focus_row_keys_from_choices(choices_path):
    if not choices_path:
        return {}
    with open(choices_path, "r", encoding="utf-8-sig") as handle:
        choices = json.load(handle)
    focus = {}
    for item in (choices.get("rowChoices") or []) + (choices.get("cellChoices") or []):
        row_key = str(item.get("rowKey") or "")
        if not row_key:
            continue
        focus.setdefault(int(item.get("sheetIndex")), set()).add(row_key)
    return focus


def load_payload(ours_path, theirs_path, max_rows, max_cols, base_path=None, choices_path=None):
    wb_base = openpyxl.load_workbook(base_path, data_only=False) if base_path else None
    wb_ours = openpyxl.load_workbook(ours_path, data_only=False)
    wb_theirs = openpyxl.load_workbook(theirs_path, data_only=False)
    try:
        focus_row_keys = focus_row_keys_from_choices(choices_path)
        sheet_count = max(len(wb_base.worksheets) if wb_base else 0, len(wb_ours.worksheets), len(wb_theirs.worksheets))
        sheets = []
        for index in range(sheet_count):
            ws_base = wb_base.worksheets[index] if wb_base and index < len(wb_base.worksheets) else None
            ws_ours = wb_ours.worksheets[index] if index < len(wb_ours.worksheets) else None
            ws_theirs = wb_theirs.worksheets[index] if index < len(wb_theirs.worksheets) else None
            sheets.append(build_sheet_payload(index, ws_ours, ws_theirs, max_rows, max_cols, ws_base, focus_row_keys.get(index + 1)))

        summary = "has_impact" if any(sheet["classification"] == "has_impact" for sheet in sheets) else "no_impact"
        structure_mismatch = any(sheet["structureMismatch"] for sheet in sheets)
        return {
            "ok": True,
            "ours": str(ours_path),
            "theirs": str(theirs_path),
            "summary": summary,
            "sheetCount": sheet_count,
            "structureMismatch": structure_mismatch,
            "sheets": sheets,
        }
    finally:
        wb_ours.close()
        wb_theirs.close()
        if wb_base:
            wb_base.close()


def ensure_same_sheet_layout(wb_ours, wb_theirs, wb_base=None):
    ours_names = [ws.title for ws in wb_ours.worksheets]
    theirs_names = [ws.title for ws in wb_theirs.worksheets]
    if ours_names != theirs_names:
        fail("sheet layout mismatch; candidate write is disabled")
    if wb_base is not None and [ws.title for ws in wb_base.worksheets] != ours_names:
        fail("base sheet layout mismatch; candidate write is disabled")


def sheet_by_index(workbook, sheet_index):
    index = int(sheet_index) - 1
    if index < 0 or index >= len(workbook.worksheets):
        fail(f"invalid sheet index: {sheet_index}")
    return workbook.worksheets[index]


def copy_cell_value(target_ws, source_ws, address):
    target_cell = target_ws[address]
    if isinstance(target_cell, MergedCell):
        return
    target_cell.value = copy.copy(source_ws[address].value)


def copy_cell_value_to(target_ws, target_address, source_ws, source_address):
    target_cell = target_ws[target_address]
    if isinstance(target_cell, MergedCell):
        return
    target_cell.value = copy.copy(source_ws[source_address].value)


def set_cell_value(target_ws, row_num, col_num, value):
    target_cell = target_ws.cell(row=row_num, column=col_num)
    if isinstance(target_cell, MergedCell):
        return
    target_cell.value = copy.copy(value)


def apply_row_choice(target_ws, source_ws, row_num):
    max_col = max(target_ws.max_column, source_ws.max_column)
    for col_num in range(1, max_col + 1):
        set_cell_value(target_ws, row_num, col_num, source_ws.cell(row=row_num, column=col_num).value)


def copy_row_values(target_ws, source_ws, source_row, target_row):
    max_col = max(target_ws.max_column, source_ws.max_column)
    for col_num in range(1, max_col + 1):
        set_cell_value(target_ws, target_row, col_num, source_ws.cell(row=source_row, column=col_num).value)


def append_row_values(target_ws, source_ws, source_row):
    target_row = target_ws.max_row + 1
    copy_row_values(target_ws, source_ws, source_row, target_row)
    return target_row


def ensure_row_at_source_position(target_ws, source_ws, source_row):
    source_row = int(source_row)
    if source_row <= target_ws.max_row:
        target_ws.insert_rows(source_row)
        target_row = source_row
    else:
        target_row = source_row
    copy_row_values(target_ws, source_ws, source_row, target_row)
    return target_row


def parse_cell_address(address):
    letters = "".join(ch for ch in str(address) if ch.isalpha())
    digits = "".join(ch for ch in str(address) if ch.isdigit())
    if not letters or not digits:
        fail(f"invalid cell address: {address}")
    return int(digits), column_index_from_string(letters.upper())


def remap_address(address, row_num):
    if not row_num:
        return ""
    _, col_num = parse_cell_address(address)
    return f"{get_column_letter(col_num)}{int(row_num)}"


def normalize_choice_action(item, expected_default="ours"):
    if item.get("action") == "keep-both":
        primary = item.get("primary") if item.get("primary") in ("ours", "theirs") else "ours"
        secondary = item.get("secondary") if item.get("secondary") in ("ours", "theirs") else ("theirs" if primary == "ours" else "ours")
        if primary == secondary:
            secondary = "theirs" if primary == "ours" else "ours"
        placement = str(item.get("placement") or "")
        if placement not in ("insert-row-after", "insert-column-after"):
            placement = "insert-column-after"
        return {
            "action": "keep-both",
            "primary": primary,
            "secondary": secondary,
            "placement": placement,
        }
    source = item.get("source") or expected_default
    if source not in ("ours", "theirs"):
        fail(f"invalid choice source: {source}")
    return {"action": "choose", "source": source}


def choice_maps(choices):
    row_choices = {}
    cell_choices = {}
    for item in choices.get("rowChoices") or []:
        row_choices[(int(item.get("sheetIndex")), int(item.get("row")))] = normalize_choice_action(item)
    for item in choices.get("cellChoices") or []:
        cell_choices[(int(item.get("sheetIndex")), str(item.get("cell")))] = normalize_choice_action(item)
    return row_choices, cell_choices


def choice_source_for_cell(row_choices, cell_choices, sheet_index, row_num, address):
    choice = selected_choice(row_choices, cell_choices, sheet_index, row_num, address)
    if not choice or choice.get("action") == "keep-both":
        return ""
    return choice.get("source") or "ours"


def selected_choice(row_choices, cell_choices, sheet_index, row_num, address):
    return cell_choices.get((sheet_index, address)) or row_choices.get((sheet_index, row_num))


def source_workbook(source, wb_ours, wb_theirs):
    return wb_theirs if source == "theirs" else wb_ours


def source_worksheet(source, wb_ours, wb_theirs, sheet_index):
    return sheet_by_index(source_workbook(source, wb_ours, wb_theirs), sheet_index)


def choice_source_row(item, source):
    if source == "ours":
        return int(item.get("oursRow") or item.get("row"))
    if source == "theirs":
        return int(item.get("theirsRow") or item.get("row"))
    return int(item.get("row"))


def choice_source_cell(item, source):
    if source == "ours":
        return item.get("oursCell") or remap_address(item.get("cell"), item.get("oursRow")) or item.get("cell")
    if source == "theirs":
        return item.get("theirsCell") or remap_address(item.get("cell"), item.get("theirsRow")) or item.get("cell")
    return item.get("cell")


def choice_target_row(item):
    value = item.get("oursRow") or item.get("row")
    return int(value) if value else 0


def choice_target_cell(item):
    return item.get("oursCell") or remap_address(item.get("cell"), item.get("oursRow")) or item.get("cell")


def choice_target_row_key(item):
    return str(item.get("rowKey") or "")


def choice_visual_row(item):
    row_num, _ = parse_cell_address(item.get("cell"))
    return int(row_num)


def keyed_sheet_context(base_ws, ours_ws, theirs_ws):
    if base_ws is None:
        return None
    base_cells, base_max_row, base_max_col = read_non_empty_cells(base_ws)
    ours_cells, ours_max_row, ours_max_col = read_non_empty_cells(ours_ws)
    theirs_cells, theirs_max_row, theirs_max_col = read_non_empty_cells(theirs_ws)
    used_cols = max(base_max_col, ours_max_col, theirs_max_col)
    if not (has_luban_header(base_cells) or has_luban_header(ours_cells) or has_luban_header(theirs_cells)):
        return None
    key_col = detect_key_column(base_cells, ours_cells, theirs_cells, used_cols)
    if not key_col:
        return None
    base_rows, base_row_numbers = row_map_by_key(base_cells, base_max_row, used_cols, key_col)
    ours_rows, ours_row_numbers = row_map_by_key(ours_cells, ours_max_row, used_cols, key_col)
    theirs_rows, theirs_row_numbers = row_map_by_key(theirs_cells, theirs_max_row, used_cols, key_col)
    keys = ordered_keys(base_rows.keys(), ours_rows.keys(), theirs_rows.keys())
    return {
        "keyCol": key_col,
        "usedCols": used_cols,
        "keys": keys,
        "baseRows": base_rows,
        "oursRows": ours_rows,
        "theirsRows": theirs_rows,
        "baseRowNumbers": base_row_numbers,
        "oursRowNumbers": ours_row_numbers,
        "theirsRowNumbers": theirs_row_numbers,
    }


def keyed_sheet_context_without_base(ours_ws, theirs_ws):
    ours_cells, ours_max_row, ours_max_col = read_non_empty_cells(ours_ws)
    theirs_cells, theirs_max_row, theirs_max_col = read_non_empty_cells(theirs_ws)
    used_cols = max(ours_max_col, theirs_max_col)
    if not (has_luban_header(ours_cells) or has_luban_header(theirs_cells)):
        return None
    key_col = detect_key_column({}, ours_cells, theirs_cells, used_cols)
    if not key_col:
        return None
    ours_rows, ours_row_numbers = row_map_by_key(ours_cells, ours_max_row, used_cols, key_col)
    theirs_rows, theirs_row_numbers = row_map_by_key(theirs_cells, theirs_max_row, used_cols, key_col)
    keys = ordered_keys((), ours_rows.keys(), theirs_rows.keys())
    return {
        "keyCol": key_col,
        "usedCols": used_cols,
        "keys": keys,
        "oursRows": ours_rows,
        "theirsRows": theirs_rows,
        "oursRowNumbers": ours_row_numbers,
        "theirsRowNumbers": theirs_row_numbers,
    }


def visual_row_for_key(context, row_key):
    try:
        return LUBAN_DATA_START_ROW + context["keys"].index(row_key)
    except ValueError:
        return None


def rebuild_target_row_numbers(target_ws, key_col, used_cols):
    cells, max_row, _ = read_non_empty_cells(target_ws)
    _, row_numbers = row_map_by_key(cells, max_row, used_cols, key_col)
    return row_numbers


def merge_keyed_sheet_with_base(target_ws, base_ws, ours_ws, theirs_ws, sheet_index, row_choices, cell_choices):
    context = keyed_sheet_context(base_ws, ours_ws, theirs_ws)
    if context is None:
        return None

    auto_merged = 0
    choice_cells = 0
    keep_both_cells = []
    keep_both_cell_rows = []
    unresolved = []
    appended_rows = 0
    deleted_rows = []
    target_row_numbers = rebuild_target_row_numbers(target_ws, context["keyCol"], context["usedCols"])

    for row_key in context["keys"]:
        visual_row = visual_row_for_key(context, row_key)
        base_exists = row_key in context["baseRows"]
        ours_exists = row_key in context["oursRows"]
        theirs_exists = row_key in context["theirsRows"]
        target_row = target_row_numbers.get(row_key)

        if not base_exists and theirs_exists and not ours_exists:
            ensure_row_at_source_position(target_ws, theirs_ws, context["theirsRowNumbers"][row_key])
            appended_rows += 1
            target_row_numbers = rebuild_target_row_numbers(target_ws, context["keyCol"], context["usedCols"])
            auto_merged += context["usedCols"]
            continue

        if base_exists and ours_exists and not theirs_exists:
            row_changed_in_ours = any(
                not values_equal(value_from_row(context["baseRows"], row_key, col_num), value_from_row(context["oursRows"], row_key, col_num))
                for col_num in range(1, context["usedCols"] + 1)
            )
            if not row_changed_in_ours and target_row:
                deleted_rows.append(target_row)
                auto_merged += context["usedCols"]
            elif row_changed_in_ours:
                unresolved.append({
                    "sheetIndex": sheet_index,
                    "cell": f"{get_column_letter(context['keyCol'])}{visual_row}",
                    "base": to_json_value(value_from_row(context["baseRows"], row_key, context["keyCol"])),
                    "ours": to_json_value(value_from_row(context["oursRows"], row_key, context["keyCol"])),
                    "theirs": None,
                })
            continue

        if not target_row or not ours_exists:
            continue

        for col_num in range(1, context["usedCols"] + 1):
            base_value = value_from_row(context["baseRows"], row_key, col_num)
            ours_value = value_from_row(context["oursRows"], row_key, col_num)
            theirs_value = value_from_row(context["theirsRows"], row_key, col_num)
            if values_equal(ours_value, theirs_value):
                continue
            ours_changed = not values_equal(base_value, ours_value)
            theirs_changed = not values_equal(base_value, theirs_value)
            address = f"{get_column_letter(col_num)}{visual_row}"
            if ours_changed and not theirs_changed:
                set_cell_value(target_ws, target_row, col_num, ours_value)
                auto_merged += 1
            elif theirs_changed and not ours_changed:
                set_cell_value(target_ws, target_row, col_num, theirs_value)
                auto_merged += 1
            elif ours_changed and theirs_changed:
                cell_choice = cell_choices.get((sheet_index, address))
                row_choice = row_choices.get((sheet_index, visual_row))
                choice = cell_choice or row_choice
                if cell_choice and cell_choice.get("action") == "keep-both":
                    primary_value = ours_value if choice["primary"] == "ours" else theirs_value
                    set_cell_value(target_ws, target_row, col_num, primary_value)
                    keep_both_item = {
                        "sheetIndex": sheet_index,
                        "row": target_row,
                        "cell": f"{get_column_letter(col_num)}{target_row}",
                        "primary": choice["primary"],
                        "secondary": choice["secondary"],
                        "placement": choice.get("placement") or "insert-column-after",
                    }
                    if keep_both_item["placement"] == "insert-row-after":
                        keep_both_cell_rows.append(keep_both_item)
                    else:
                        keep_both_cells.append(keep_both_item)
                    choice_cells += 1
                elif row_choice and row_choice.get("action") == "keep-both":
                    primary_value = ours_value if row_choice["primary"] == "ours" else theirs_value
                    set_cell_value(target_ws, target_row, col_num, primary_value)
                    choice_cells += 1
                elif choice and choice.get("source") == "ours":
                    set_cell_value(target_ws, target_row, col_num, ours_value)
                    choice_cells += 1
                elif choice and choice.get("source") == "theirs":
                    set_cell_value(target_ws, target_row, col_num, theirs_value)
                    choice_cells += 1
                else:
                    unresolved.append({
                        "sheetIndex": sheet_index,
                        "cell": address,
                        "base": to_json_value(base_value),
                        "ours": to_json_value(ours_value),
                        "theirs": to_json_value(theirs_value),
                    })

    for row_num in sorted(set(deleted_rows), reverse=True):
        target_ws.delete_rows(row_num)

    return {
        "autoMerged": auto_merged,
        "unresolved": unresolved,
        "choiceCells": choice_cells,
        "keepBothCells": keep_both_cells,
        "keepBothCellRows": keep_both_cell_rows,
        "keyedSheets": 1,
        "keyedRowsAppended": appended_rows,
        "keyedRowsDeleted": len(set(deleted_rows)),
    }


def merge_cells_with_base(wb_target, wb_base, wb_ours, wb_theirs, row_choices, cell_choices):
    if wb_base is None:
        return {"autoMerged": 0, "unresolved": [], "choiceCells": 0, "keepBothCells": [], "keepBothCellRows": [], "keyedSheets": 0, "keyedRowsAppended": 0, "keyedRowsDeleted": 0}
    auto_merged = 0
    choice_cells = 0
    keep_both_cells = []
    keep_both_cell_rows = []
    unresolved = []
    keyed_sheets = 0
    keyed_rows_appended = 0
    keyed_rows_deleted = 0
    for index, target_ws in enumerate(wb_target.worksheets):
        sheet_index = index + 1
        base_ws = wb_base.worksheets[index]
        ours_ws = wb_ours.worksheets[index]
        theirs_ws = wb_theirs.worksheets[index]
        keyed_result = merge_keyed_sheet_with_base(target_ws, base_ws, ours_ws, theirs_ws, sheet_index, row_choices, cell_choices)
        if keyed_result is not None:
            auto_merged += keyed_result["autoMerged"]
            choice_cells += keyed_result["choiceCells"]
            keep_both_cells.extend(keyed_result["keepBothCells"])
            keep_both_cell_rows.extend(keyed_result["keepBothCellRows"])
            unresolved.extend(keyed_result["unresolved"])
            keyed_sheets += keyed_result["keyedSheets"]
            keyed_rows_appended += keyed_result["keyedRowsAppended"]
            keyed_rows_deleted += keyed_result["keyedRowsDeleted"]
            continue
        max_row = max(base_ws.max_row, ours_ws.max_row, theirs_ws.max_row)
        max_col = max(base_ws.max_column, ours_ws.max_column, theirs_ws.max_column)
        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_col + 1):
                base_value = base_ws.cell(row=row_num, column=col_num).value
                ours_value = ours_ws.cell(row=row_num, column=col_num).value
                theirs_value = theirs_ws.cell(row=row_num, column=col_num).value
                if values_equal(ours_value, theirs_value):
                    continue
                address = f"{get_column_letter(col_num)}{row_num}"
                ours_changed = not values_equal(base_value, ours_value)
                theirs_changed = not values_equal(base_value, theirs_value)
                if ours_changed and not theirs_changed:
                    set_cell_value(target_ws, row_num, col_num, ours_value)
                    auto_merged += 1
                elif theirs_changed and not ours_changed:
                    set_cell_value(target_ws, row_num, col_num, theirs_value)
                    auto_merged += 1
                elif ours_changed and theirs_changed:
                    cell_choice = cell_choices.get((sheet_index, address))
                    row_choice = row_choices.get((sheet_index, row_num))
                    choice = cell_choice or row_choice
                    if cell_choice and cell_choice.get("action") == "keep-both":
                        primary_value = ours_value if choice["primary"] == "ours" else theirs_value
                        set_cell_value(target_ws, row_num, col_num, primary_value)
                        keep_both_item = {
                            "sheetIndex": sheet_index,
                            "row": row_num,
                            "cell": address,
                            "primary": choice["primary"],
                            "secondary": choice["secondary"],
                            "placement": choice.get("placement") or "insert-column-after",
                        }
                        if keep_both_item["placement"] == "insert-row-after":
                            keep_both_cell_rows.append(keep_both_item)
                        else:
                            keep_both_cells.append(keep_both_item)
                        choice_cells += 1
                    elif row_choice and row_choice.get("action") == "keep-both":
                        primary_value = ours_value if row_choice["primary"] == "ours" else theirs_value
                        set_cell_value(target_ws, row_num, col_num, primary_value)
                        choice_cells += 1
                    elif choice and choice.get("source") == "ours":
                        set_cell_value(target_ws, row_num, col_num, ours_value)
                        choice_cells += 1
                    elif choice and choice.get("source") == "theirs":
                        set_cell_value(target_ws, row_num, col_num, theirs_value)
                        choice_cells += 1
                    else:
                        unresolved.append({
                            "sheetIndex": sheet_index,
                            "cell": address,
                            "base": to_json_value(base_value),
                            "ours": to_json_value(ours_value),
                        "theirs": to_json_value(theirs_value),
                        })
    return {
        "autoMerged": auto_merged,
        "unresolved": unresolved,
        "choiceCells": choice_cells,
        "keepBothCells": keep_both_cells,
        "keepBothCellRows": keep_both_cell_rows,
        "keyedSheets": keyed_sheets,
        "keyedRowsAppended": keyed_rows_appended,
        "keyedRowsDeleted": keyed_rows_deleted,
    }


def unresolved_cells_without_base(wb_ours, wb_theirs, row_choices, cell_choices):
    unresolved = []
    for index, ours_ws in enumerate(wb_ours.worksheets):
        sheet_index = index + 1
        theirs_ws = wb_theirs.worksheets[index]
        ours_cells, ours_max_row, ours_max_col = read_non_empty_cells(ours_ws)
        theirs_cells, theirs_max_row, theirs_max_col = read_non_empty_cells(theirs_ws)
        used_cols = max(ours_max_col, theirs_max_col)
        key_col = detect_key_column({}, ours_cells, theirs_cells, used_cols)
        if key_col and (has_luban_header(ours_cells) or has_luban_header(theirs_cells)):
            ours_rows, _ = row_map_by_key(ours_cells, ours_max_row, used_cols, key_col)
            theirs_rows, _ = row_map_by_key(theirs_cells, theirs_max_row, used_cols, key_col)
            keys = ordered_keys((), ours_rows.keys(), theirs_rows.keys())
            for visual_offset, row_key in enumerate(keys):
                visual_row = LUBAN_DATA_START_ROW + visual_offset
                for col_num in range(1, used_cols + 1):
                    ours_value = value_from_row(ours_rows, row_key, col_num)
                    theirs_value = value_from_row(theirs_rows, row_key, col_num)
                    if values_equal(ours_value, theirs_value):
                        continue
                    address = f"{get_column_letter(col_num)}{visual_row}"
                    if choice_source_for_cell(row_choices, cell_choices, sheet_index, visual_row, address):
                        continue
                    unresolved.append({
                        "sheetIndex": sheet_index,
                        "cell": address,
                        "ours": to_json_value(ours_value),
                        "theirs": to_json_value(theirs_value),
                    })
            continue

        max_row = max(ours_ws.max_row, theirs_ws.max_row)
        max_col = max(ours_ws.max_column, theirs_ws.max_column)
        for row_num in range(1, max_row + 1):
            for col_num in range(1, max_col + 1):
                ours_value = ours_ws.cell(row=row_num, column=col_num).value
                theirs_value = theirs_ws.cell(row=row_num, column=col_num).value
                if values_equal(ours_value, theirs_value):
                    continue
                address = f"{get_column_letter(col_num)}{row_num}"
                if choice_source_for_cell(row_choices, cell_choices, sheet_index, row_num, address):
                    continue
                unresolved.append({
                    "sheetIndex": sheet_index,
                    "cell": address,
                    "ours": to_json_value(ours_value),
                    "theirs": to_json_value(theirs_value),
                })
    return unresolved


def apply_keyed_choices_without_base(wb_target, wb_ours, wb_theirs, cell_choices):
    applied = 0
    appended_rows = 0
    handled_cells = set()
    inserted_rows = {}
    by_sheet = {}
    for item in cell_choices:
        if normalize_choice_action(item).get("action") == "keep-both":
            continue
        by_sheet.setdefault(int(item.get("sheetIndex")), []).append(item)

    for sheet_index, items in by_sheet.items():
        target_ws = sheet_by_index(wb_target, sheet_index)
        ours_ws = sheet_by_index(wb_ours, sheet_index)
        theirs_ws = sheet_by_index(wb_theirs, sheet_index)
        context = keyed_sheet_context_without_base(ours_ws, theirs_ws)
        if context is None:
            continue

        target_row_numbers = rebuild_target_row_numbers(target_ws, context["keyCol"], context["usedCols"])
        appended_keys = set()
        rows_to_insert = []
        for item in items:
            choice = normalize_choice_action(item)
            if choice.get("source") != "theirs":
                continue
            row_key = str(item.get("rowKey") or "")
            if not row_key or row_key in target_row_numbers or row_key in appended_keys:
                continue
            source_row = context["theirsRowNumbers"].get(row_key)
            if not source_row:
                continue
            rows_to_insert.append((source_row, row_key))
            appended_keys.add(row_key)

        for source_row, row_key in sorted(rows_to_insert):
            inserted_rows[(sheet_index, row_key)] = ensure_row_at_source_position(target_ws, theirs_ws, source_row)
            appended_rows += 1

        if appended_keys:
            target_row_numbers = rebuild_target_row_numbers(target_ws, context["keyCol"], context["usedCols"])

        for item in items:
            choice = normalize_choice_action(item)
            if choice.get("action") == "keep-both":
                continue
            row_key = str(item.get("rowKey") or "")
            target_row = inserted_rows.get((sheet_index, row_key)) or target_row_numbers.get(row_key)
            if not target_row:
                continue
            _, col_num = parse_cell_address(item.get("cell"))
            source = choice.get("source") or "ours"
            source_row = context[f"{source}RowNumbers"].get(row_key)
            if not source_row:
                continue
            source_ws = ours_ws if source == "ours" else theirs_ws
            set_cell_value(target_ws, target_row, col_num, source_ws.cell(row=source_row, column=col_num).value)
            applied += 1
            handled_cells.add((sheet_index, str(item.get("cell"))))

    return {"cellChoices": applied, "keyedRowsAppended": appended_rows, "handledCells": handled_cells}


def apply_keyed_row_choice_without_base(wb_target, wb_ours, wb_theirs, item, choice):
    row_key = choice_target_row_key(item)
    if not row_key:
        return False
    sheet_index = int(item.get("sheetIndex"))
    target_ws = sheet_by_index(wb_target, sheet_index)
    ours_ws = sheet_by_index(wb_ours, sheet_index)
    theirs_ws = sheet_by_index(wb_theirs, sheet_index)
    context = keyed_sheet_context_without_base(ours_ws, theirs_ws)
    if context is None:
        return False

    target_row_numbers = rebuild_target_row_numbers(target_ws, context["keyCol"], context["usedCols"])
    source = choice.get("source") or "ours"
    source_row = context[f"{source}RowNumbers"].get(row_key)
    if not source_row:
        return True

    target_row = target_row_numbers.get(row_key)
    if target_row:
        copy_row_values(target_ws, ours_ws if source == "ours" else theirs_ws, source_row, target_row)
        return True

    if source == "theirs":
        ensure_row_at_source_position(target_ws, theirs_ws, source_row)
    return True


def row_choice_sort_key(item):
    choice = normalize_choice_action(item)
    if choice.get("action") == "keep-both":
        return (int(item.get("sheetIndex")), int(item.get("row") or 0))
    source = choice.get("source") or "ours"
    return (int(item.get("sheetIndex")), choice_source_row(item, source))


def apply_keep_both_rows(wb_target, wb_ours, wb_theirs, row_choices):
    applied = 0
    for (sheet_index, row_num), choice in sorted(row_choices.items(), key=lambda item: (item[0][0], -item[0][1])):
        if choice.get("action") != "keep-both":
            continue
        target_ws = sheet_by_index(wb_target, sheet_index)
        primary_ws = source_worksheet(choice["primary"], wb_ours, wb_theirs, sheet_index)
        secondary_ws = source_worksheet(choice["secondary"], wb_ours, wb_theirs, sheet_index)
        copy_row_values(target_ws, primary_ws, row_num, row_num)
        target_ws.insert_rows(row_num + 1)
        copy_row_values(target_ws, secondary_ws, row_num, row_num + 1)
        applied += 1
    return applied


def apply_keep_both_cells(wb_target, wb_ours, wb_theirs, keep_both_cells):
    applied = 0
    groups = {}
    for item in keep_both_cells:
        row_num, col_num = parse_cell_address(item["cell"])
        groups.setdefault((item["sheetIndex"], col_num), []).append((row_num, item))

    for (sheet_index, col_num), cells in sorted(groups.items(), key=lambda value: (value[0][0], -value[0][1])):
        target_ws = sheet_by_index(wb_target, sheet_index)
        target_ws.insert_cols(col_num + 1)
        target_ws.cell(row=1, column=col_num + 1).value = target_ws.cell(row=1, column=col_num + 1).value or f"{get_column_letter(col_num)}_THEIRS"
        for row_num, item in cells:
            primary_ws = source_worksheet(item["primary"], wb_ours, wb_theirs, sheet_index)
            secondary_ws = source_worksheet(item["secondary"], wb_ours, wb_theirs, sheet_index)
            set_cell_value(target_ws, row_num, col_num, primary_ws.cell(row=row_num, column=col_num).value)
            set_cell_value(target_ws, row_num, col_num + 1, secondary_ws.cell(row=row_num, column=col_num).value)
            applied += 1
    return applied


def apply_keep_both_cell_rows(wb_target, wb_ours, wb_theirs, keep_both_cell_rows):
    applied = 0
    groups = {}
    for item in keep_both_cell_rows:
        groups.setdefault((item["sheetIndex"], item["row"], item["primary"]), []).append(item)

    for (sheet_index, row_num, primary), cells in sorted(groups.items(), key=lambda item: (item[0][0], -item[0][1])):
        target_ws = sheet_by_index(wb_target, sheet_index)
        primary_ws = source_worksheet(primary, wb_ours, wb_theirs, sheet_index)
        copy_row_values(target_ws, primary_ws, row_num, row_num)
        target_ws.insert_rows(row_num + 1)
        copy_row_values(target_ws, primary_ws, row_num, row_num + 1)
        for item in cells:
            _, col_num = parse_cell_address(item["cell"])
            secondary_ws = source_worksheet(item["secondary"], wb_ours, wb_theirs, sheet_index)
            set_cell_value(target_ws, row_num + 1, col_num, secondary_ws.cell(row=row_num, column=col_num).value)
            applied += 1
    return applied


def write_candidate(ours_path, theirs_path, choices_path, output_path, base_path=None):
    wb_base = openpyxl.load_workbook(base_path, data_only=False) if base_path else None
    wb_target = openpyxl.load_workbook(ours_path, data_only=False)
    wb_ours_source = openpyxl.load_workbook(ours_path, data_only=False)
    wb_theirs = openpyxl.load_workbook(theirs_path, data_only=False)
    try:
        ensure_same_sheet_layout(wb_target, wb_theirs, wb_base)
        with open(choices_path, "r", encoding="utf-8-sig") as handle:
            choices = json.load(handle)

        row_choices = choices.get("rowChoices") or []
        cell_choices = choices.get("cellChoices") or []
        row_choice_map, cell_choice_map = choice_maps(choices)

        row_applied = 0
        cell_applied = 0
        auto_merge = merge_cells_with_base(wb_target, wb_base, wb_ours_source, wb_theirs, row_choice_map, cell_choice_map)
        keyed_no_base = {"cellChoices": 0, "keyedRowsAppended": 0, "handledCells": set()}
        if wb_base is None:
            keyed_no_base = apply_keyed_choices_without_base(wb_target, wb_ours_source, wb_theirs, cell_choices)
            auto_merge["unresolved"] = unresolved_cells_without_base(wb_ours_source, wb_theirs, row_choice_map, cell_choice_map)
        if auto_merge["unresolved"]:
            cells = ", ".join(f"Sheet{item['sheetIndex']}:{item['cell']}" for item in auto_merge["unresolved"][:20])
            suffix = " ..." if len(auto_merge["unresolved"]) > 20 else ""
            fail(f"unresolved Excel cell conflicts require explicit choices: {cells}{suffix}")
        for item in sorted(row_choices, key=row_choice_sort_key):
            choice = normalize_choice_action(item)
            if choice["action"] == "keep-both":
                row_applied += 1
                continue
            source = choice["source"]
            if wb_base is not None:
                row_applied += 1
                continue
            if apply_keyed_row_choice_without_base(wb_target, wb_ours_source, wb_theirs, item, choice):
                row_applied += 1
                continue
            target_ws = sheet_by_index(wb_target, item.get("sheetIndex"))
            source_wb = wb_ours_source if source == "ours" else wb_theirs
            source_ws = sheet_by_index(source_wb, item.get("sheetIndex"))
            copy_row_values(target_ws, source_ws, choice_source_row(item, source), choice_target_row(item))
            row_applied += 1

        for item in cell_choices:
            choice = normalize_choice_action(item)
            if choice["action"] == "keep-both":
                cell_applied += 1
                continue
            if (int(item.get("sheetIndex")), str(item.get("cell"))) in keyed_no_base["handledCells"]:
                cell_applied += 1
                continue
            source = choice["source"]
            if wb_base is not None:
                cell_applied += 1
                continue
            target_ws = sheet_by_index(wb_target, item.get("sheetIndex"))
            source_wb = wb_ours_source if source == "ours" else wb_theirs
            source_ws = sheet_by_index(source_wb, item.get("sheetIndex"))
            copy_cell_value_to(target_ws, choice_target_cell(item), source_ws, choice_source_cell(item, source))
            cell_applied += 1

        keep_both_rows = apply_keep_both_rows(wb_target, wb_ours_source, wb_theirs, row_choice_map)
        keep_both_cell_rows = apply_keep_both_cell_rows(wb_target, wb_ours_source, wb_theirs, auto_merge["keepBothCellRows"])
        keep_both_cells = apply_keep_both_cells(wb_target, wb_ours_source, wb_theirs, auto_merge["keepBothCells"])

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb_target.save(output_path)
        return {
            "ok": True,
            "output": str(output_path),
            "applied": {
                "rowChoices": row_applied,
                "cellChoices": cell_applied,
                "choiceCells": auto_merge["choiceCells"] + keyed_no_base["cellChoices"],
                "autoMergedCells": auto_merge["autoMerged"],
                "keepBothRows": keep_both_rows,
                "keepBothCellRows": keep_both_cell_rows,
                "keepBothCells": keep_both_cells,
                "keyedSheets": auto_merge.get("keyedSheets", 0),
                "keyedRowsAppended": auto_merge.get("keyedRowsAppended", 0) + keyed_no_base["keyedRowsAppended"],
                "keyedRowsDeleted": auto_merge.get("keyedRowsDeleted", 0),
            },
            "unresolved": auto_merge["unresolved"],
        }
    finally:
        if wb_base:
            wb_base.close()
        wb_target.close()
        wb_ours_source.close()
        wb_theirs.close()


def main():
    parser = argparse.ArgumentParser(description="Excel conflict workbench helper")
    parser.add_argument("--mode", choices=("load", "write-candidate"), required=True)
    parser.add_argument("--ours", required=True)
    parser.add_argument("--theirs", required=True)
    parser.add_argument("--base")
    parser.add_argument("--choices")
    parser.add_argument("--output")
    parser.add_argument("--max-rows", type=int, default=MAX_ROWS_DEFAULT)
    parser.add_argument("--max-cols", type=int, default=MAX_COLS_DEFAULT)
    args = parser.parse_args()

    ours_path = Path(args.ours)
    theirs_path = Path(args.theirs)
    if not ours_path.exists():
        fail(f"OURS workbook does not exist: {ours_path}")
    if not theirs_path.exists():
        fail(f"THEIRS workbook does not exist: {theirs_path}")
    if not is_xlsx(ours_path) or not is_xlsx(theirs_path):
        fail("only .xlsx workbooks are supported")
    base_path = Path(args.base) if args.base else None
    if base_path and not base_path.exists():
        fail(f"BASE workbook does not exist: {base_path}")
    if base_path and not is_xlsx(base_path):
        fail("base workbook must be .xlsx")

    if args.mode == "load":
        result = load_payload(ours_path, theirs_path, max(1, args.max_rows), max(1, args.max_cols), base_path, args.choices)
    else:
        if not args.choices:
            fail("--choices is required for write-candidate")
        if not args.output:
            fail("--output is required for write-candidate")
        if not is_xlsx(args.output):
            fail("candidate output must be .xlsx")
        result = write_candidate(ours_path, theirs_path, args.choices, args.output, base_path)

    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
